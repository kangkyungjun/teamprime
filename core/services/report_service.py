"""
리포트 자동 생성 서비스
- Excel 리포트 생성 (openpyxl)
- PDF 리포트 생성 (reportlab)
- 차트 및 그래프 포함
- 템플릿 기반 리포트
"""

import logging
import io
import os
import base64
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any, BinaryIO
from dataclasses import dataclass
import json
import math

logger = logging.getLogger(__name__)

@dataclass
class ReportData:
    """리포트 데이터 컨테이너"""
    title: str
    period: str
    generated_at: str
    business_performance: Dict
    task_efficiency: Dict
    expense_analysis: Dict
    predictions: List[Dict]
    summary: Dict

class ReportService:
    """리포트 생성 서비스"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        
    async def generate_excel_report(
        self, 
        report_data: ReportData, 
        report_type: str = "comprehensive"
    ) -> io.BytesIO:
        """Excel 리포트 생성"""
        
        try:
            # openpyxl 동적 import (설치되어 있지 않을 경우 대비)
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.chart import LineChart, BarChart, PieChart, Reference
                from openpyxl.utils.dataframe import dataframe_to_rows
            except ImportError:
                return await self._generate_csv_fallback(report_data)
            
            # 워크북 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "종합 리포트"
            
            # 스타일 정의
            header_font = Font(name='맑은 고딕', size=14, bold=True)
            title_font = Font(name='맑은 고딕', size=16, bold=True)
            normal_font = Font(name='맑은 고딕', size=11)
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            light_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            
            center_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'), 
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 1. 리포트 헤더
            ws.merge_cells('A1:H1')
            ws['A1'] = f"업무 관리 시스템 - {report_data.title}"
            ws['A1'].font = title_font
            ws['A1'].alignment = center_alignment
            
            ws.merge_cells('A2:H2')
            ws['A2'] = f"기간: {report_data.period} | 생성일시: {report_data.generated_at}"
            ws['A2'].font = normal_font
            ws['A2'].alignment = center_alignment
            
            current_row = 4
            
            # 2. 요약 정보
            ws.merge_cells(f'A{current_row}:H{current_row}')
            ws[f'A{current_row}'] = "📊 핵심 지표 요약"
            ws[f'A{current_row}'].font = header_font
            ws[f'A{current_row}'].fill = header_fill
            current_row += 1
            
            summary = report_data.summary
            summary_data = [
                ["총 수익", f"{summary.get('total_income', 0):,}원"],
                ["총 지출", f"{summary.get('total_expense', 0):,}원"],
                ["순이익", f"{summary.get('net_profit', 0):,}원"],
                ["완료된 업무", f"{summary.get('completed_tasks', 0)}건"],
                ["업무 완료율", f"{summary.get('completion_rate', 0):.1f}%"],
                ["평균 업무 처리시간", f"{summary.get('avg_processing_time', 0):.1f}시간"]
            ]
            
            for row_data in summary_data:
                ws[f'A{current_row}'] = row_data[0]
                ws[f'B{current_row}'] = row_data[1]
                ws[f'A{current_row}'].font = normal_font
                ws[f'B{current_row}'].font = normal_font
                current_row += 1
            
            current_row += 2
            
            # 3. 비즈니스 성과 분석
            ws.merge_cells(f'A{current_row}:H{current_row}')
            ws[f'A{current_row}'] = "💰 비즈니스 성과 분석"
            ws[f'A{current_row}'].font = header_font
            ws[f'A{current_row}'].fill = header_fill
            current_row += 1
            
            # 월별 데이터
            if 'monthly_data' in report_data.business_performance:
                ws[f'A{current_row}'] = "월"
                ws[f'B{current_row}'] = "수익"
                ws[f'C{current_row}'] = "지출"
                ws[f'D{current_row}'] = "순이익"
                
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{current_row}'].font = header_font
                    ws[f'{col}{current_row}'].fill = light_fill
                    
                current_row += 1
                
                for month_data in report_data.business_performance['monthly_data']:
                    ws[f'A{current_row}'] = month_data.get('month', '')
                    ws[f'B{current_row}'] = month_data.get('income', 0)
                    ws[f'C{current_row}'] = month_data.get('expense', 0)
                    ws[f'D{current_row}'] = month_data.get('profit', 0)
                    current_row += 1
            
            current_row += 2
            
            # 4. 업무 효율성 분석
            ws.merge_cells(f'A{current_row}:H{current_row}')
            ws[f'A{current_row}'] = "⚡ 업무 효율성 분석"
            ws[f'A{current_row}'].font = header_font
            ws[f'A{current_row}'].fill = header_fill
            current_row += 1
            
            efficiency = report_data.task_efficiency
            efficiency_data = [
                ["전체 업무 수", efficiency.get('total_tasks', 0)],
                ["완료된 업무", efficiency.get('completed_tasks', 0)],
                ["진행 중인 업무", efficiency.get('in_progress_tasks', 0)],
                ["지연된 업무", efficiency.get('overdue_tasks', 0)],
                ["평균 완료 시간", f"{efficiency.get('avg_completion_time', 0):.1f}시간"],
                ["생산성 점수", f"{efficiency.get('productivity_score', 0):.1f}점"]
            ]
            
            for row_data in efficiency_data:
                ws[f'A{current_row}'] = row_data[0]
                ws[f'B{current_row}'] = row_data[1]
                current_row += 1
            
            current_row += 2
            
            # 5. 예측 정보
            if report_data.predictions:
                ws.merge_cells(f'A{current_row}:H{current_row}')
                ws[f'A{current_row}'] = "🔮 향후 6개월 예측"
                ws[f'A{current_row}'].font = header_font
                ws[f'A{current_row}'].fill = header_fill
                current_row += 1
                
                ws[f'A{current_row}'] = "월"
                ws[f'B{current_row}'] = "예상 수익"
                ws[f'C{current_row}'] = "예상 지출"
                ws[f'D{current_row}'] = "예상 순이익"
                ws[f'E{current_row}'] = "신뢰구간 하한"
                ws[f'F{current_row}'] = "신뢰구간 상한"
                
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    ws[f'{col}{current_row}'].font = header_font
                    ws[f'{col}{current_row}'].fill = light_fill
                    
                current_row += 1
                
                for pred in report_data.predictions[:6]:  # 6개월만
                    ws[f'A{current_row}'] = pred.get('date', '')
                    ws[f'B{current_row}'] = pred.get('predicted_income', 0)
                    ws[f'C{current_row}'] = pred.get('predicted_expense', 0)
                    ws[f'D{current_row}'] = pred.get('predicted_profit', 0)
                    ws[f'E{current_row}'] = pred.get('confidence_lower', 0)
                    ws[f'F{current_row}'] = pred.get('confidence_upper', 0)
                    current_row += 1
            
            # 파일을 BytesIO로 저장
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            self.logger.info(f"Excel 리포트 생성 완료: {report_data.title}")
            return excel_file
            
        except Exception as e:
            self.logger.error(f"Excel 리포트 생성 실패: {str(e)}")
            return await self._generate_csv_fallback(report_data)
    
    async def _generate_csv_fallback(self, report_data: ReportData) -> io.BytesIO:
        """CSV 대체 리포트 생성"""
        import csv
        
        csv_content = io.StringIO()
        writer = csv.writer(csv_content)
        
        # 헤더
        writer.writerow([f"업무 관리 시스템 - {report_data.title}"])
        writer.writerow([f"기간: {report_data.period} | 생성일시: {report_data.generated_at}"])
        writer.writerow([])
        
        # 요약 정보
        writer.writerow(["핵심 지표 요약"])
        summary = report_data.summary
        writer.writerow(["항목", "값"])
        writer.writerow(["총 수익", f"{summary.get('total_income', 0):,}원"])
        writer.writerow(["총 지출", f"{summary.get('total_expense', 0):,}원"])
        writer.writerow(["순이익", f"{summary.get('net_profit', 0):,}원"])
        writer.writerow([])
        
        # BytesIO로 변환
        csv_bytes = io.BytesIO()
        csv_bytes.write(csv_content.getvalue().encode('utf-8-sig'))
        csv_bytes.seek(0)
        
        return csv_bytes
    
    async def generate_pdf_report(
        self, 
        report_data: ReportData, 
        include_charts: bool = True
    ) -> io.BytesIO:
        """PDF 리포트 생성"""
        
        try:
            # reportlab 동적 import
            try:
                from reportlab.lib.pagesizes import A4, letter
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib import colors
                from reportlab.lib.units import inch
                from reportlab.pdfgen import canvas
                from reportlab.pdfbase import pdfmetrics
                from reportlab.pdfbase.ttfonts import TTFont
            except ImportError:
                return await self._generate_text_fallback(report_data)
            
            # PDF 문서 생성
            pdf_buffer = io.BytesIO()
            doc = SimpleDocTemplate(pdf_buffer, pagesize=A4)
            
            # 스타일 설정
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1  # 중앙 정렬
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=14,
                spaceAfter=12
            )
            
            normal_style = styles['Normal']
            
            # 문서 내용
            story = []
            
            # 제목
            title = Paragraph(f"업무 관리 시스템 - {report_data.title}", title_style)
            story.append(title)
            
            subtitle = Paragraph(f"기간: {report_data.period}<br/>생성일시: {report_data.generated_at}", normal_style)
            story.append(subtitle)
            story.append(Spacer(1, 20))
            
            # 요약 정보
            story.append(Paragraph("📊 핵심 지표 요약", heading_style))
            
            summary = report_data.summary
            summary_data = [
                ['항목', '값'],
                ['총 수익', f"{summary.get('total_income', 0):,}원"],
                ['총 지출', f"{summary.get('total_expense', 0):,}원"],
                ['순이익', f"{summary.get('net_profit', 0):,}원"],
                ['완료된 업무', f"{summary.get('completed_tasks', 0)}건"],
                ['업무 완료율', f"{summary.get('completion_rate', 0):.1f}%"],
                ['평균 업무 처리시간', f"{summary.get('avg_processing_time', 0):.1f}시간"]
            ]
            
            summary_table = Table(summary_data)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(summary_table)
            story.append(Spacer(1, 20))
            
            # 비즈니스 성과 분석
            story.append(Paragraph("💰 비즈니스 성과 분석", heading_style))
            
            if 'monthly_data' in report_data.business_performance:
                monthly_data = [['월', '수익', '지출', '순이익']]
                for month_data in report_data.business_performance['monthly_data'][:12]:  # 최대 12개월
                    monthly_data.append([
                        month_data.get('month', ''),
                        f"{month_data.get('income', 0):,}원",
                        f"{month_data.get('expense', 0):,}원",
                        f"{month_data.get('profit', 0):,}원"
                    ])
                
                monthly_table = Table(monthly_data)
                monthly_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(monthly_table)
                story.append(Spacer(1, 20))
            
            # 업무 효율성 분석
            story.append(Paragraph("⚡ 업무 효율성 분석", heading_style))
            
            efficiency = report_data.task_efficiency
            efficiency_data = [
                ['지표', '값'],
                ['전체 업무 수', str(efficiency.get('total_tasks', 0))],
                ['완료된 업무', str(efficiency.get('completed_tasks', 0))],
                ['진행 중인 업무', str(efficiency.get('in_progress_tasks', 0))],
                ['지연된 업무', str(efficiency.get('overdue_tasks', 0))],
                ['평균 완료 시간', f"{efficiency.get('avg_completion_time', 0):.1f}시간"],
                ['생산성 점수', f"{efficiency.get('productivity_score', 0):.1f}점"]
            ]
            
            efficiency_table = Table(efficiency_data)
            efficiency_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(efficiency_table)
            story.append(Spacer(1, 20))
            
            # 예측 정보
            if report_data.predictions:
                story.append(Paragraph("🔮 향후 6개월 예측", heading_style))
                
                prediction_data = [['월', '예상 수익', '예상 지출', '예상 순이익']]
                for pred in report_data.predictions[:6]:  # 6개월만
                    prediction_data.append([
                        pred.get('date', ''),
                        f"{pred.get('predicted_income', 0):,.0f}원",
                        f"{pred.get('predicted_expense', 0):,.0f}원",
                        f"{pred.get('predicted_profit', 0):,.0f}원"
                    ])
                
                prediction_table = Table(prediction_data)
                prediction_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(prediction_table)
            
            # PDF 생성
            doc.build(story)
            pdf_buffer.seek(0)
            
            self.logger.info(f"PDF 리포트 생성 완료: {report_data.title}")
            return pdf_buffer
            
        except Exception as e:
            self.logger.error(f"PDF 리포트 생성 실패: {str(e)}")
            return await self._generate_text_fallback(report_data)
    
    async def _generate_text_fallback(self, report_data: ReportData) -> io.BytesIO:
        """텍스트 대체 리포트 생성"""
        
        text_content = f"""
업무 관리 시스템 - {report_data.title}
기간: {report_data.period}
생성일시: {report_data.generated_at}

==========================================
📊 핵심 지표 요약
==========================================

총 수익: {report_data.summary.get('total_income', 0):,}원
총 지출: {report_data.summary.get('total_expense', 0):,}원
순이익: {report_data.summary.get('net_profit', 0):,}원
완료된 업무: {report_data.summary.get('completed_tasks', 0)}건
업무 완료율: {report_data.summary.get('completion_rate', 0):.1f}%
평균 업무 처리시간: {report_data.summary.get('avg_processing_time', 0):.1f}시간

==========================================
💰 비즈니스 성과 분석
==========================================

"""
        
        if 'monthly_data' in report_data.business_performance:
            text_content += "월별 성과:\n"
            for month_data in report_data.business_performance['monthly_data'][:12]:
                text_content += f"{month_data.get('month', '')}: 수익 {month_data.get('income', 0):,}원, "
                text_content += f"지출 {month_data.get('expense', 0):,}원, "
                text_content += f"순이익 {month_data.get('profit', 0):,}원\n"
        
        text_content += f"""
==========================================
⚡ 업무 효율성 분석
==========================================

전체 업무 수: {report_data.task_efficiency.get('total_tasks', 0)}
완료된 업무: {report_data.task_efficiency.get('completed_tasks', 0)}
진행 중인 업무: {report_data.task_efficiency.get('in_progress_tasks', 0)}
지연된 업무: {report_data.task_efficiency.get('overdue_tasks', 0)}
평균 완료 시간: {report_data.task_efficiency.get('avg_completion_time', 0):.1f}시간
생산성 점수: {report_data.task_efficiency.get('productivity_score', 0):.1f}점
"""
        
        if report_data.predictions:
            text_content += "\n==========================================\n"
            text_content += "🔮 향후 6개월 예측\n"
            text_content += "==========================================\n\n"
            
            for pred in report_data.predictions[:6]:
                text_content += f"{pred.get('date', '')}: "
                text_content += f"예상 수익 {pred.get('predicted_income', 0):,.0f}원, "
                text_content += f"예상 지출 {pred.get('predicted_expense', 0):,.0f}원, "
                text_content += f"예상 순이익 {pred.get('predicted_profit', 0):,.0f}원\n"
        
        # BytesIO로 변환
        text_bytes = io.BytesIO()
        text_bytes.write(text_content.encode('utf-8'))
        text_bytes.seek(0)
        
        return text_bytes
    
    async def generate_chart_image(self, chart_data: Dict, chart_type: str) -> Optional[bytes]:
        """차트 이미지 생성 (matplotlib 사용)"""
        
        try:
            import matplotlib
            matplotlib.use('Agg')  # GUI 없는 환경에서 사용
            import matplotlib.pyplot as plt
            from matplotlib import font_manager
            import numpy as np
            
            # 한글 폰트 설정
            plt.rcParams['font.family'] = ['DejaVu Sans', 'Arial Unicode MS', 'sans-serif']
            plt.rcParams['axes.unicode_minus'] = False
            
            fig, ax = plt.subplots(figsize=(10, 6))
            
            if chart_type == 'line':
                # 선 그래프
                x_data = chart_data.get('labels', [])
                y_data = chart_data.get('data', [])
                ax.plot(x_data, y_data, marker='o', linewidth=2, markersize=6)
                ax.set_title(chart_data.get('title', ''), fontsize=14, fontweight='bold')
                ax.grid(True, alpha=0.3)
                
            elif chart_type == 'bar':
                # 막대 그래프
                x_data = chart_data.get('labels', [])
                y_data = chart_data.get('data', [])
                bars = ax.bar(x_data, y_data, color='skyblue', alpha=0.8)
                ax.set_title(chart_data.get('title', ''), fontsize=14, fontweight='bold')
                
                # 막대 위에 값 표시
                for bar in bars:
                    height = bar.get_height()
                    ax.text(bar.get_x() + bar.get_width()/2., height,
                           f'{height:,.0f}',
                           ha='center', va='bottom')
                           
            elif chart_type == 'pie':
                # 파이 차트
                labels = chart_data.get('labels', [])
                data = chart_data.get('data', [])
                colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#FFA07A', '#98D8C8']
                
                wedges, texts, autotexts = ax.pie(data, labels=labels, colors=colors[:len(data)], 
                                                 autopct='%1.1f%%', startangle=90)
                ax.set_title(chart_data.get('title', ''), fontsize=14, fontweight='bold')
            
            plt.tight_layout()
            
            # 이미지를 바이트로 변환
            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format='PNG', dpi=300, bbox_inches='tight')
            img_buffer.seek(0)
            img_bytes = img_buffer.read()
            plt.close()
            
            return img_bytes
            
        except Exception as e:
            self.logger.warning(f"차트 이미지 생성 실패: {str(e)}")
            return None
    
    def get_report_filename(self, report_type: str, file_format: str, period: str = None) -> str:
        """리포트 파일명 생성"""
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        period_str = f"_{period}" if period else ""
        
        return f"업무관리_리포트_{report_type}{period_str}_{timestamp}.{file_format.lower()}"

# 전역 리포트 서비스 인스턴스
report_service = ReportService()